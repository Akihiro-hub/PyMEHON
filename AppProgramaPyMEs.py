import streamlit as st

import time
def main():
    while True:
        time.sleep(360 * 360)  

import pandas as pd
from io import BytesIO
from openpyxl import Workbook
import matplotlib.pyplot as plt
import numpy as np
import xgboost as xgb
from sklearn.model_selection import train_test_split
from sklearn.metrics import mean_squared_error

from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment, Border, Side  # 必要なモジュールをインポート

from wordcloud import WordCloud
import seaborn as sns
from collections import Counter
import PyPDF2
import re

rubro = st.sidebar.selectbox("Herramientas de planificación a aplicar", ["Seleccione", "Plan de negocio en operación",  "Pronóstico de ventas", "Simulación de inversión", "Estudio de mercadeo", "Planificación de inventario", "Planificación de venta (Comedor)"])

if rubro == "Seleccione":
    st.write("## Aplicación digital para la planificación y gestión de negocio de PyMEs")
    st.write("### :blue[(Programa PyMEs del Proyecto EDIFICA)]")
    st.write("##### Esta aplicación contiene diferentes herramientas para facilitar la planificación y gestión de negocio de PyMEs. Dichas herramientas incluyen;") 
    st.write("(A) Plan de negocio en operación")
    st.write("(B) Simulación de proyecto de inversión")
    st.write("(C) Pronósitico de ventas")
    st.write("(D) Estudio de mercadeo")
    st.write("(E) Planificación de inventario, etc")

    st.write("###### (NOTA: Cada una de herramientas se presentará, dependiendo de su selección en las opciones presentadas a la izquierda.)")

elif rubro == "Pronóstico de ventas":
    # 過去12か月の売上データの初期値
    ventas_iniciales = [21000, 17500, 18000, 18500, 25000, 21000, 19000, 22000, 23500, 19500, 21000, 23000]
    # 過去12か月のその他の特徴量（Touristasは2022年の数値、千人単位、Cruceristas（通過者）は含まない。家族送金は2003～23年の月間平均。単位100万ドル）
    turistas = [44, 51, 71, 86, 69, 81, 85, 75, 54, 63, 71, 96]
    remesas = [477, 488, 581, 572, 618, 623, 606, 633, 599, 636, 573, 641]
    
    st.write("### :blue[Pronóstico (estimación) de ventas en próximos 12 meses]")
    st.write("###### (Herramienta de Inteligencia Artificial por Modelo XGBoost, ajustado del método de los mínimos cuadrados, para sectores de comercio y turísmo)")
    st.write("###### :red[Esta herramienta estima las ventas en futuro próximo, mediante la información sobre las ventas realizadas en estos 12 meses, los datos climáticos de la ciudad (a seleccionar) y el monto de remesas familiares por mes, el número de visitantes exteriores al país. Será probable que el resultado de estimación no sea precisa, debido a la limitación de los datos de variables explicativas.]")
    
    # 各都市のデータ
    ciudades = {
        "Tegucigalpa": {
            "lluvias": [0.4, 0.5, 0.5, 2.7, 9.8, 13.3, 10.6, 12.3, 15.0, 11.1, 3.7, 1.3],
            "temperaturas": [20, 21, 22, 24, 24, 23, 22, 23, 22, 22, 21, 20],
        },
        "San Marcos de Colón": {
            "lluvias": [0.2, 0.3, 0.8, 2.4, 8.9, 11.7, 8.5, 11.0, 13.5, 10.9, 3.5, 1.0],
            "temperaturas": [21, 22, 23, 24, 23, 22, 22, 22, 21, 21, 21, 21],
        },
        "Choluteca": {
            "lluvias": [0.1, 0.2, 0.7, 2.3, 8.9, 11.7, 8.6, 11.0, 13.6, 10.7, 3.2, 0.9],
            "temperaturas": [29, 29, 30, 31, 29, 28, 29, 29, 27, 27, 28, 29],
        },
        "Santa Rosa de Copán": {
            "lluvias": [1.9, 1.8, 1.6, 3.0, 8.2, 13.2, 12.3, 12.3, 13.4, 9.9, 4.8, 2.9],
            "temperaturas": [18, 19, 20, 22, 22, 22, 21, 22, 21, 20, 19, 18],
        },
        "San Pedro Sula": {
            "lluvias": [5.9, 4.7, 3.8, 3.4, 6.4, 11.1, 11.3, 11.2, 11.5, 10.3, 8.5, 7.2],
            "temperaturas": [23, 24, 25, 27, 28, 28, 27, 27, 27, 26, 24, 24],
        },
    }
    # 月の選択肢
    meses = ["Ene", "Feb", "Mar", "Abr", "May", "Jun", "Jul", "Ago", "Sep", "Oct", "Nov", "Dic"]
    
    st.write("##### :blue[Seleccione el mes actual y la ciudad cuyo clima es semejante al mismo de su lugar]")
    
    col1, col2 = st.columns(2)
    with col1:
        # 選択された月の初期値
        mes_actual = st.selectbox("Selecciona el mes actual", meses, index=7)
    
    with col2:
        # Select the city
        ciudad = st.selectbox("Selecciona la ciudad", list(ciudades.keys()))
    
    # Get the city's data
    lluvias = ciudades[ciudad]["lluvias"]
    temperaturas = ciudades[ciudad]["temperaturas"]
    
    # 月のインデックスを取得
    mes_index = meses.index(mes_actual)
    
    # ユーザーが売上データを入力
    st.write("##### :blue[Ingrese los datos de ventas de los últimos 12 meses]")
    
    # 各列に4か月分の売上データ入力フィールドを配置するための列の作成
    cols = st.columns(4)
    
    # 12か月前からの順序を保持し、各列に4か月分を表示
    for i in range(12):
        col_index = i // 3  # 0, 1, 2, 3 (4列)
        month_label = f"Hace {12 - i} meses ({meses[(mes_index - 12 + i) % 12]})"
        with cols[col_index]:
            ventas_iniciales[i] = st.number_input(month_label, value=ventas_iniciales[i], key=i)
    
    # データフレームの作成
    data = pd.DataFrame({
        'Ventas': ventas_iniciales,
        "Días de lluvias": lluvias[mes_index:] + lluvias[:mes_index],
        "Temperatura mínima del día": temperaturas[mes_index:] + temperaturas[:mes_index],
        'Visitantes exteriores al país': turistas[mes_index:] + turistas[:mes_index],
        "Remesas familiares": remesas[mes_index:] + remesas[:mes_index],
    })
    
    # 特徴量とターゲットの準備
    X = data[['Días de lluvias', 'Temperatura mínima del día', 'Visitantes exteriores al país', 'Remesas familiares']]
    y = data['Ventas']
    
    # データを訓練セットとテストセットに分割
    X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=0.3, shuffle=False)
    
    # XGBoostモデルの訓練
    model = xgb.XGBRegressor(objective='reg:squarederror', n_estimators=13)
    model.fit(X_train, y_train)
    
    # 12カ月先まで予測
    forecast_input = X.iloc[-1].values.reshape(1, -1)
    forecast = []
    for i in range(12):
        next_pred = model.predict(forecast_input)[0]
        forecast.append(next_pred)
        # 新しい特徴量の生成
        new_row = np.array([lluvias[(mes_index + i + 1) % 12], temperaturas[(mes_index + i + 1) % 12], turistas[(mes_index + i + 1) % 12], remesas[(mes_index + i + 1) % 12]]).reshape(1, -1)
        forecast_input = new_row
    
    forecast_df = pd.DataFrame(forecast, index=[f"{meses[(mes_index+i)%12]}" for i in range(12, 24)], columns=['Ventas'])
    forecast_df['Ventas'] = forecast_df['Ventas'].round(0).astype(int)  # 売上高を整数に丸める

    # 最小二乗法で傾きを計算
    from scipy.stats import linregress
    x = np.arange(len(ventas_iniciales))
    slope, intercept, _, _, _ = linregress(x, ventas_iniciales)

    # 傾きを加算して予測を修正
    forecast_df['Ventas'] = forecast_df['Ventas'] + slope * np.arange(1, 13)
    
    # 実績データと予測データの結合
    full_data = pd.concat([data, forecast_df])
    full_data.index = [f"Hace {12-i} meses ({meses[(mes_index-12+i)%12]})" for i in range(12)] + [meses[(mes_index+i)%12] for i in range(12, 24)]
    
    if st.button("Estimar (pronosticar) ventas futuras por la inteligencia artificial"):
    
        # グラフの表示
        st.subheader("Ventas realizadas y estimadas en los 24 meses")
        plt.figure(figsize=(12, 4))
        plt.plot(full_data.index[:12], full_data['Ventas'][:12], label='Ventas realizadas', color='blue', marker='o')
        plt.plot(full_data.index[12:], full_data['Ventas'][12:], label='Ventas estimadas', color='orange', marker='o')
        plt.xticks(rotation=45, ha='right')
        plt.legend(loc='upper left')
        plt.grid(True)
        plt.tight_layout()
        st.pyplot(plt)
    
        # 表の表示
        st.subheader("Datos de ventas realizadas y estimadas")
        st.write("Los datos de días de lluvia y otros indicadores no son exactamente del año pasado sino de los otros años de muestra.")
        resultados = pd.concat([data, forecast_df])
        resultados.index = [f"Hace {12-i} meses ({meses[(mes_index-12+i)%12]})" for i in range(12)] + [meses[(mes_index+i)%12] for i in range(12, 24)]
        st.dataframe(resultados)
    
        # エクセルファイルのダウンロード
        st.subheader("Descargar Datos en Excel")
        def convert_df(df):
            return df.to_csv().encode('utf-8')
        csv = convert_df(resultados)
        st.download_button(label="Descargar datos en Excel como CSV", data=csv, file_name='prediccion_ventas.csv', mime='text/csv')
    
        
elif rubro == "Plan de negocio en operación":
    st.write("## :blue[Plan de negocio en operación]") 
    st.write("###### Esta herramienta facilita la planificación del monto a vender y el flujo de caja.") 
    
    def calculate_cash_flow(initial_cash, sales, material_cost, labor_cost, loan_repayment, other_fixed_costs, desired_profit):
        fixed_cost = labor_cost + loan_repayment + other_fixed_costs
        variable_ratio = material_cost / sales
        breakeven_sales = fixed_cost / (1 - variable_ratio)
        required_sales = (fixed_cost + desired_profit) / (1 - variable_ratio)
        
        cash_flow = {
            "Saldo del ejecutivo al inicio": [],
            "Ingresos (Caja de entradas)": [],
            "Egresos (Caja de salidas)": [],
            "Saldo al final": []
        }
        for month in range(12):
            cash_inflow = sales
            cash_outflow = material_cost + labor_cost + loan_repayment + other_fixed_costs
            month_end_cash = initial_cash + cash_inflow - cash_outflow
            cash_flow["Saldo del ejecutivo al inicio"].append(initial_cash)
            cash_flow["Ingresos (Caja de entradas)"].append(cash_inflow)
            cash_flow["Egresos (Caja de salidas)"].append(cash_outflow)
            cash_flow["Saldo al final"].append(month_end_cash)
            initial_cash = month_end_cash
        return breakeven_sales, required_sales, cash_flow, fixed_cost, variable_ratio

    def generate_excel(cash_flow):
        wb = Workbook()
        ws = wb.active
        ws.title = "Presupuesto del flujo de caja"

        headers = ["", "1r mes", "2do mes", "3r mes", "4to mes", "5to mes", "6to mes", "7mo mes", "8vo mes", "9no mes", "10mo mes", "11 mes", "12 mes"]
        for col_num, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_num, value=header)

        for row_num, (key, values) in enumerate(cash_flow.items(), 2):
            ws.cell(row=row_num, column=1, value=key)
            for col_num, value in enumerate(values, 2):
                ws.cell(row=row_num, column=col_num, value=value)

        excel_data = BytesIO()
        wb.save(excel_data)
        excel_data.seek(0)
        
        return excel_data

    col1, col2 = st.columns(2)
    with col1:
        sales = st.number_input("Monto estimado de venta mensual (¿Cuánto monto su negocio vende al mes en Lps?):", min_value=0, value=16000, step=1, format="%d")
        desired_profit = st.number_input("Meta de ganancias mensuales (¿Cuánto desea ganar al mes en Lps?):", min_value=0, value=5000, step=1, format="%d")
        initial_cash = st.number_input("Saldo inicial del ejecutivo (¿Cuánto monto de ejecutivo comercial tiene actualmente en Lps?):", min_value=0, value=4500, step=1, format="%d")
    with col2:
        material_cost = st.number_input("Costo mensual de materias primas (y otros costos variables, Lps):", min_value=0, value=6000, step=1, format="%d")
        labor_cost = st.number_input("Remuneraciones mensuales de trabajadores como costo fijo (Lps):", min_value=0, value=4000, step=1, format="%d")
        loan_repayment = st.number_input("Pago mensual de deuda (como costo fijo, Lps):", min_value=0, value=0, step=1, format="%d")
        other_fixed_costs = st.number_input("Otros costos fijos, tales como alquiler de la tienda, electricidad, etc (Lps):", min_value=0, value=4500, step=1, format="%d")
       
    if st.button("Elaborar el plan operativo de negocio (planificación de venta y flujo de caja)"):
        breakeven_sales, required_sales, cash_flow, fixed_cost, variable_ratio = calculate_cash_flow(
            initial_cash, sales, material_cost, labor_cost, loan_repayment, other_fixed_costs, desired_profit)

        months = ["1r mes", "2do mes", "3r mes", "4to mes", "5to mes", "6to mes", "7mo mes", "8vo mes", "9no mes", "10mo mes", "11 mes", "12 mes"]
        df = pd.DataFrame(cash_flow, index=months).T
        
        st.write("#### :blue[(1) Planificación de ventas, en base al análisis del punto de equilibrio]") 
        st.write(f"Ventas al mes en el punto de equilibrio: {breakeven_sales:.2f} Lps")
        st.write(f"Ventas necesarias para lograr la meta de ganancias al mes: {required_sales:.2f} Lps")

        fig, ax = plt.subplots()
        
        sales_range = list(range(int(breakeven_sales * 0.8), int(required_sales * 1.2), 100))
        total_costs = [fixed_cost + (variable_ratio * s) for s in sales_range]
        
        ax.plot(sales_range, total_costs, color='skyblue', label="Costos totales (Costos fijos + Costos variables)", marker='o')
        ax.plot(sales_range, sales_range, color='orange', label="Venta", marker='o')
        
        ax.set_title("Análisis de punto de equilibrio")
        ax.set_xlabel("Venta (Lps)")
        ax.set_ylabel("Costos y ventas (Lps)")
        
        ax.axvline(breakeven_sales, color='red', linestyle='--', label=f"Punto de equilibrio: {breakeven_sales:.2f} Lps")
        
        ax.fill_between(sales_range, total_costs, sales_range, where=[s > breakeven_sales for s in sales_range], color='skyblue', alpha=0.3, interpolate=True)
        
        mid_x = (required_sales + breakeven_sales) / 2
        mid_y = (max(total_costs) + max(sales_range)) / 2
        ax.text(mid_x, mid_y, "Ganancia = Área del color azul claro", color="blue", fontsize=7, ha="center")

        ax.legend()  # Show the legend
        st.pyplot(fig)
        
        st.write("#### :blue[(2) Presupuesto del flujo de caja por 12 meses]") 
        st.dataframe(df)

        excel_data = generate_excel(cash_flow)
        st.download_button(
            label="Descargar la tabla EXCEL",
            data=excel_data,
            file_name="business_plan.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        st.write("###### Puede descargar la tabla en Excel. Es recomendable elaborar el plan del flujo de caja de manera más precisa, aplicando la otra herramienta, puesto que la tabla presentada arriba es de versión muy resumida.") 


elif rubro == "Planificación de inventario":
    st.write("## :blue[Planificación de inventario de seguridad]") 
    st.write("###### Esta herramienta facilita la calculación del volumen de inventario de seguridad, que se refiere a la cantidad necesaria a mantener siempre para evitar escasez, en ciertas materias importantes.")  
    st.write("###### En el siguiente ejemplo se muestra un caso de maderas aserradas como la materia prima principal de la carpintería, mientras que esta herramienta es aplicable para otros negocios también.")
    st.write("###### Es importante calcular el volumen de inventario de seguridad, ya que el mismo se relaciona directamente al monto necesario del capital de trabajo.")
    col1, col2 = st.columns(2)
    with col1:
        a = st.number_input("¿Hace 5 días (o semana) cuántas piezas de madera aserrada se consumieron?", 0, 10000, 30)
        b = st.number_input("¿Hace 4 días (o semana) cuántas piezas de madera aserrada se consumieron?", 0, 10000, 25)
        c = st.number_input("¿Hace 3 días (o semana) cuántas piezas de madera aserrada se consumieron?", 0, 10000, 45)
    with col2:
        d = st.number_input("¿Hace 2 días (o semana) cuántas piezas de madera aserrada se consumieron?", 0, 10000, 37)
        e = st.number_input("¿Ayer (o semana pasada) cuántas piezas de madera aserrada se consumieron?", 0, 10000, 18)
        g = st.number_input("¿Cuánto días (o semanas) debe esperar la recepción de maderas después de la colocación de la orden?", 0, 300, 5)
    data = [a, b, c, d, e]
    SD = np.std(data, ddof=1) 
    import math
    Inventario_seguridad1 = 2.33 * SD * math.sqrt(g)
    Inventario_seguridad5 = 1.64 * SD * math.sqrt(g)   
    Inventario_seguridad10 = 1.28 * SD * math.sqrt(g)

    if st.button("Calcular el volumen de inventario de seguridad)"):
        st.write("##### Resultado de cálculo:") 
        col1, col2 = st.columns(2)
        with col1:
            st.write("##### :green[Volumen de inventario de seguridad]")
            st.write("###### Caso A: Inventario de seguridad con la probabilidad de escasez de 1% (piezas):")
            st.text(round(Inventario_seguridad1))
            st.write("###### Caso B: Inventario de seguridad con la probabilidad de escasez de 5% (piezas):")
            st.text(round(Inventario_seguridad5))
            st.write("###### Caso C: Inventario de seguridad con la probabilidad de escasez de 10% (piezas):")
            st.text(round(Inventario_seguridad10))  
        with col2:
            st.write("##### :green[Volumen al punto de ordenar ]")
            st.write("###### Volumen de inventario en posesión al punto de ordenar en Caso A (piezas):")
            st.text(round(Inventario_seguridad1+np.mean(data)*g))
            st.write("###### Volumen de inventario en posesión al punto de ordenar en Caso B (piezas):")
            st.text(round(Inventario_seguridad5+np.mean(data)*g))
            st.write("###### Volumen de inventario en posesión al punto de ordenar en Caso C (piezas):")
            st.text(round(Inventario_seguridad10+np.mean(data)*g))  
        st.write("###### :red[NOTA: Además del inventario de seguridad, la empresa también necesita tener cierto volumen del inventario para su consumo durante el período de espera después de colocación de la orden de materias primas, por lo que el volumen de inventario a tener al punto de ordenar debe ser mayor que el inventario de seguridad. En otras palabras, el volumen al punto de colocación de la orden puede ser; Promedio de consumos diarios x Días de espera + Inventario de seguridad.]")

elif rubro == "Planificación de venta (Comedor)":
    st.write("## :blue[Planificación del monto de ventas en un comedor]") 
    st.write("###### El monto de la venta de un restaurante, comedor o cafetería se puede estimar, en base al número de asientos, aplicando esta calculadora.")  
   
    col1, col2 = st.columns(2)
    with col1:
        a = st.number_input("¿Cuánto asientos tiene el comedor?", 0, 1000, 20)
        b = st.number_input("Tasa de ocupación de los asientos por los clientes (%)", 0, 100, 50)
        c = st.number_input("Veces estimadas de rotación de los clientes al día", 1, 10, 3)
    with col2:
        d = st.number_input("Promedio estimado de la venta por cliente (Lps)", 1, 1000, 125)
        e = st.number_input("Días de operación al mes (Días)", 1, 31, 28)
    st.write("###### :red[La tasa de ocupación puede ser 50%, ya que sólo dos personas pueden ocupar la mesa para cuatro personas. La rotacion de los clientes al día puede ser 4 o 5 veces, como 2 rotaciones a horas de almuerzo y 2 rotaciones a horas de cena.]")
    
    E = a*d*(b/100)*c

    if st.button("Estimar el monto de ventas"):
        st.write("##### Resultado del cálculo: Monto esperado de la venta diaria")
        st.text(E)
        st.write("##### Resultado del cálculo: Monto esperado de la venta mensual")
        st.text(E*e)


elif rubro == "Simulación de inversión":
    # 設定: アプリケーションの基本設定
    st.title("Simulación de Proyecto de Inversión por PyME")
    st.write("###### :blue[Esta herramienta facilita una simulación sencilla de inversión por PyMEs para maquinaria o nuevo negocio. Es recomendable realizar los estudios más detallados, al concretar el plan de inversión.]") 
    st.write("###### Ingrese los datos principales del proyecto de inversión a analizar.")

    # 初期値の設定
    # Col1, Col2, Col3の作成（間にスペースを挟む）
    col1, col2_space, col2, col3_space, col3 = st.columns([0.9, 0.05, 0.8, 0.05, 1.25])
    with col1: 
        st.write("###### :red[Maquinaria o Equipo a comprar (O, inversión para el nuevo negocio):]")
        inversion_inicial = st.number_input("Monto de inversión (Lps)", value=100000)
        vida_util = st.number_input("Años de vida útil del equipo a invertir", value=6)
        st.write("###### :red[Tasa de impuesto:]")
        tasa_impuesto = st.number_input("Tasa de impuesto (%)", value=15)

    with col2:
        st.write("###### :red[Posible Uso de Crédito:]")
        monto_prestamo = st.number_input("Monto del préstamo a aplicar para la inversión (Lps)", value=60000)
        tasa_interes = st.number_input("Tasa de interés del préstamo (%)", value=20)
        meses_prestamo = st.number_input("Meses de reembolso del préstamo", value=30)
        
    with col3:
        st.write("###### :red[Ingresos y costos del proyecto:]")
        ventas_anuales = st.number_input("Ventas anuales (adicionales) a generar por el proyecto (Lps)", value=90000)
        costos_ventas = st.number_input("Proporción (%) de costos productivos sobre las ventas (Nota: Los costos productivos son de materias primas, trabajadores productivos, y otros relacionados al proyecto, excluyendo depreciación)", value=60)
        gastos_administrativos = st.number_input("Gastos administrativos anuales relacionado al proyecto (Lps)", value=5000)

    # Analizarボタンの設定
    if st.button("Analizar"):

        # 償還表作成
        # 月利の計算
        monthly_rate = tasa_interes / 100 / 12
        
        # 月数に基づいて毎月の返済額の計算
        monthly_payment = (monto_prestamo * monthly_rate * (1 + monthly_rate) ** meses_prestamo) / ((1 + monthly_rate) ** meses_prestamo - 1)

        # 初期設定
        balance = monto_prestamo
        schedule = []

        # 各月の償還表を作成
        for month in range(1, meses_prestamo + 1):
            interest_payment = balance * monthly_rate
            principal_payment = monthly_payment - interest_payment
            balance -= principal_payment
            schedule.append([month, round(monthly_payment), round(principal_payment), round(interest_payment), round(balance)])

        # データフレームに変換し、インデックスを表示しない
        df = pd.DataFrame(schedule, columns=["Mes", "Pago mensual (Lps)", "Pago a capital (Lps)", "Interés (Lps)", "Saldo restante (Lps)"])

        st.subheader("A) Cuadro de Amortización del crédito a solicitar en base al plan de cuotas niveladas")
        st.dataframe(df.reset_index(drop=True))

        # 年間利息支払額の計算
        df['Año'] = (df['Mes'] - 1) // 12 + 1  # 各行に対応する年を計算
        yearly_interest = df.groupby('Año')['Interés (Lps)'].sum().reset_index()  # 年ごとの利息の合計を計算

        # 年間元本支払額の計算
        yearly_capital = df.groupby('Año')['Pago a capital (Lps)'].sum().reset_index()

        # 月単位の調整
        full_years = meses_prestamo // 12
        remaining_months = meses_prestamo % 12

        # 年ごとの金利負担
        intereses = yearly_interest['Interés (Lps)'].tolist()[:full_years]

        # 端数調整（月単位での追加）
        if remaining_months > 0:
            third_year_interest = df[df['Año'] == full_years + 1]['Interés (Lps)'].sum()
            intereses.append(third_year_interest)

        # 金利負担のない月はゼロ表示
        intereses += [0] * (vida_util - full_years - 1)

        # 予想損益計算書の作成
        st.subheader("B) Estado de Resultados Proyectado")
        ventas = np.array([ventas_anuales] * vida_util)
        costo_ventas_sin_depreciacion = ventas * (costos_ventas / 100)
        depreciacion = inversion_inicial / vida_util
        costo_total_ventas = costo_ventas_sin_depreciacion + depreciacion
        utilidad_bruta = ventas - costo_total_ventas
        utilidad_operativa = utilidad_bruta - gastos_administrativos

        # 確実に配列として扱うために、interesesをNumPy配列に変換
        intereses = np.array(intereses)

        # 利益に関する計算もNumPy配列として扱う
        utilidad_operativa = np.array(utilidad_operativa)

        # 税前利益・純利益の計算
        utilidad_antes_impuestos = utilidad_operativa - intereses  # 配列同士の引き算
        utilidad_neta = utilidad_antes_impuestos * (1 - tasa_impuesto/100)

        # 損益計算書のデータフレーム
        data_sonkei = {
            "Año": list(range(1, vida_util + 1)),
            "Ventas": ventas,
            "Costos productivos": costo_ventas_sin_depreciacion,
            "Depreciación": [depreciacion] * vida_util,
            "Costo total de ventas": costo_total_ventas,
            "Utilidad bruta": utilidad_bruta,
            "Gastos administrativos": [gastos_administrativos] * vida_util,
            "Intereses": intereses,
            "Utilidad antes de impuestos": utilidad_antes_impuestos,
            "Utilidad neta": utilidad_neta,
        }
        df_sonkei = pd.DataFrame(data_sonkei).T.round(0)  # 小数点以下を四捨五入して整数表示
        st.dataframe(df_sonkei)

        # キャッシュフロー計算書
        st.subheader("C) Estado de Flujo de Caja Proyectado")

        # flujo_operativoの定義（vida_utilの年数に合わせる）現在はゼロ
        flujo_operativo = [0] + list(utilidad_neta + depreciacion)

        # flujo_inversionの定義（vida_utilに合わせる、最初の年に-inversion_inicial、それ以外は0）
        flujo_inversion = [-inversion_inicial] + [0] * vida_util

        # flujo_financieroの定義（年数をvida_utilに合わせる）
        flujo_financiero = [monto_prestamo] + [-capital for capital in yearly_capital['Pago a capital (Lps)']]
        
        # flujo_operativoがvida_utilに合うように長さを調整
        if len(flujo_operativo) < vida_util + 1:
            flujo_operativo += [0] * (vida_util + 1 - len(flujo_operativo))

        # flujo_inversionの定義
        flujo_inversion = [-inversion_inicial] + [0] * vida_util

        # flujo_financieroの長さをvida_utilに合わせる
        if len(flujo_financiero) < vida_util + 1:
            flujo_financiero += [0] * (vida_util + 1 - len(flujo_financiero))

        # 各リストの長さを確認
        print(f"flujo_operativo: {len(flujo_operativo)}")
        print(f"flujo_inversion: {len(flujo_inversion)}")
        print(f"flujo_financiero: {len(flujo_financiero)}")

        # リストの長さが一致していることを確認
        assert len(flujo_operativo) == len(flujo_inversion) == len(flujo_financiero), "リストの長さが一致していません"

        # flujo_totalの計算
        flujo_total = [flujo_operativo[i] + flujo_inversion[i] + flujo_financiero[i] for i in range(len(flujo_operativo))]

        # キャッシュフローのデータフレーム
        data_cf = {
            "Año": ["Hoy"] + list(range(1, vida_util + 1)),
            "Flujo operativo": flujo_operativo,
            "Flujo de inversión": flujo_inversion,
            "Flujo financiero": flujo_financiero,
            "Flujo neto": flujo_total
        }

        df_cf = pd.DataFrame(data_cf)

        # 数値カラムだけを整数に変換
        numeric_cols = ["Flujo operativo", "Flujo de inversión", "Flujo financiero", "Flujo neto"]
        df_cf[numeric_cols] = df_cf[numeric_cols].round(0).astype(int)

        # データフレームの転置
        df_cf_transposed = df_cf.T

        # 転置後にヘッダー行を設定
        df_cf_transposed.columns = df_cf_transposed.iloc[0]
        df_cf_transposed = df_cf_transposed[1:]

        # データフレームの表示
        st.dataframe(df_cf_transposed)

        # 投資プロジェクト評価指標の作成
        st.subheader("D) Indicadores de Evaluación del Proyecto")
        flujo_operativoOR = list(utilidad_neta + depreciacion)
        flujo_descuento = flujo_operativoOR / ((1 + tasa_interes / 100) ** np.arange(1, vida_util + 1))
        npv = np.sum(flujo_descuento) - inversion_inicial
        roi = np.sum(utilidad_antes_impuestos) / inversion_inicial
        
        rate = tasa_interes/100
        payback = 1/rate - (1/(rate*(1+rate)**vida_util))
        st.write(f"###### Valor Presente Neto (VPN): {npv:.2f} Lps")
        st.write(f"###### Rentabilidad sobre la Inversión (ROI): {roi:.1f} %")
        st.write(f"###### Periodo máximo aceptable para recaudación del fondo invertido: {payback*12:.1f} meses")
    
        st.write("###### :red[Un proyecto con el VPN negativo o insuficiente se debe rechazar. Para simplificar el calculo del VPN, se aplica la tasa de interes, como la tasa de descuentos. El tercer indicador es para la referencia teórica, y el empresario deberá recuperar el fondo invertido lo antes posible. Se presenta abajo una figura del flujo neto de caja del Proyecto.]") 

        # 棒グラフの作成
        fig, ax = plt.subplots()
        ax.bar(range(vida_util + 1), flujo_total, label='Flujo neto', color='blue')

        # X軸に年ごとのラベルを追加
        ax.set_xticks(range(vida_util + 1))
        ax.set_xticklabels([f'Año {i}' for i in range(vida_util + 1)])

        # 金額ゼロのところに水平線を追加
        ax.axhline(0, color='red', linewidth=1.5)

        # グラフのラベルとタイトル
        ax.set_xlabel('Año')
        ax.set_ylabel('Flujo de caja (Lps)')
        ax.set_title('Proyección de Flujo de caja durante el Proyecto')

        # グラフをStreamlitで表示
        st.pyplot(fig)

        # 損益分岐点分析グラフ
        st.subheader("E) Gráfico de Análisis de Punto de Equilibrio al año")
        st.write("Se presentan abajo el resultado del análisis del punto de equilibrio. El resultado del análisis podrá ser impreciso, en los siguientes dos sentidos. `Primero, en esta simulacion, los costos se clasifican en los fijos y variables de manera no precisa. Segundo, este análisis no incluye el cálculo de descuentos basado en la teoría financiera. Es decir, considerando el costo de adquisición del capital, el punto de equilibrio, en términos reales, podrá ser más alto que la cifra indicada abajo.")

        # 固定費と変動費の計算
        fixedcost = gastos_administrativos + np.mean(intereses) + inversion_inicial/vida_util # 固定費は管理費と平均利息を加えたもの
        variable_ratio = costos_ventas/100  # 変動費率を計算

        # 損益分岐点の計算
        breakeven_sales = fixedcost / (1 - variable_ratio)

        # グラフの作成
        fig, ax = plt.subplots()

        # 損益分岐点前後の売上範囲を設定
        sales_range = np.arange(int(breakeven_sales * 0.8), int(breakeven_sales * 1.2), 100)

        # 総コストを計算
        total_costs = [fixedcost + (variable_ratio * s) for s in sales_range]

        # 総コストと売上のプロット
        ax.plot(sales_range, total_costs, color='skyblue', label="Costos totales", marker='o')
        ax.plot(sales_range, sales_range, color='orange', label="Venta anual", marker='o')

        # グラフのタイトルとラベル
        ax.set_title("Estimación del punto de equilibrio")
        ax.set_xlabel("Venta (Lps)")
        ax.set_ylabel("Costos y ventas (Lps)")

        # 損益分岐点の縦線を追加
        ax.axvline(breakeven_sales, color='red', linestyle='--', label=f"Punto de equilibrio: {breakeven_sales:.0f} Lps")

        # 損益分岐点の説明
        ax.fill_between(sales_range, total_costs, sales_range, where=[s > breakeven_sales for s in sales_range], color='skyblue', alpha=0.3, interpolate=True)

        # グラフに説明を追加
        mid_x = breakeven_sales * 1.05  # 説明テキストの位置調整
        mid_y = (max(total_costs) + max(sales_range)) / 2
        ax.text(mid_x, mid_y, "Ganancia = Área del color azul claro", color="blue", fontsize=7, ha="left")

        # 凡例の表示
        ax.legend()

        # グラフをStreamlitに表示
        st.pyplot(fig)

elif rubro == "Estudio de mercadeo":
    # Streamlit UIの設定
    st.write("## :blue[Análisis de Texto o Documento]") 
    st.write("Puede elaborar, para el estudio de mercadeo u otros motivos, Nube de Palabras (WordCloud), figura visual donde las palabras frecuentes o importantes en un texto se presentan de manera destacada. Se usa en muchos proyectos de Inteligencia Artificial.")
    st.write("##### :green[Paso 1: Pegue el texto para el análisis o suba un archivo PDF.]")

    # テキストエリアとPDFファイルアップロードのオプション
    col1, col2 = st.columns(2)
    with col1:
        texto = st.text_area("Pegue aquí el texto a analizar", "")
    with col2: 
        pdf_file = st.file_uploader("O, sube un archivo PDF a analizar", type="pdf")

    # 除外したい単語の入力
    st.write("##### :green[Paso 2: Note abajo las palabras que deben excluirse del análisis, si las tiene.]")
    col1, col2, col3 = st.columns(3)
    with col1:
        exclude_word1 = st.text_input("Palabra a excluir 1", "")
        exclude_word2 = st.text_input("Palabra a excluir 2", "")
    with col2:
        exclude_word3 = st.text_input("Palabra a excluir 3", "")
        exclude_word4 = st.text_input("Palabra a excluir 4", "")
    with col3:
        exclude_word5 = st.text_input("Palabra a excluir 5", "")
        exclude_word6 = st.text_input("Palabra a excluir 6", "")

    # デフォルトで除外する単語のリスト
    default_excluded_words = {
        "la", "el", "los", "las", "él", "ella", "en", "de", "del", "un", "que", "soy", "eres", "es", "somos", "son", "inglés", "traducido",
        "estoy", "estás", "le", "poder", "hace", "año", "mes", "he", "estado", "había", "años", "meses", "sobre", "dueño", 
        "gusta", "me", "mi", "su", "opiniones", "sugerencias", "calificación", "respuesta", "propietario", "dueño", "negocio",
        "está", "estamos", "están", "este", "aquello", "aquella", "esta", "estas", "estos", "cual", "y", "ya", "hay", "a", "al",
        "lo", "desde", "hasta", "hacia", "usted", "tú", "yo", "compartir", "con", "para", "su", "nuestro", "sea", "sean", "esté", "estén", 
        "o", "u", "e", "por", "eso", "foto", "fotos", "local", "reseñas", "más", "mas", "nos", "os", "ser", "estar", "sí", "si", "estuviese", "estuviera",
        "no", "ni", "guide", "hay", "se", "una", "uno", "fuí", "fue", "fuera", "fuese", "hubiera", "estaba", "estaban", "estuve", "estuvo", "estuvieron"
    }

    # 入力された除外単語をセットに追加
    user_excluded_words = {exclude_word1, exclude_word2, exclude_word3, exclude_word4, exclude_word5, exclude_word6}
    excluded_words = default_excluded_words.union(user_excluded_words)

    if pdf_file:
        # PDFからテキストを抽出
        pdf_reader = PyPDF2.PdfReader(pdf_file)
        texto = ""
        for page in pdf_reader.pages:
            texto += page.extract_text()

    if texto:
        # テキストのトークン化と前処理
        words = re.findall(r'\b\w+\b', texto.lower())
        filtered_words = [word for word in words if word not in excluded_words]

        # Word Cloudの作成
        wordcloud = WordCloud(width=800, height=400, background_color='white').generate(" ".join(filtered_words))

    # 分析ボタンの表示
    if st.button("Analizar"):

        # Word Cloudの表示
        st.subheader("Nube de Palabras (WordCloud)")
        st.write("WordCloud es útil para visualizar la importancia relativa de términos dentro de un texto de manera rápida y comprensible.")

        fig, ax = plt.subplots()
        ax.imshow(wordcloud, interpolation='bilinear')
        ax.axis("off")
        st.pyplot(fig)
        
        # 頻出する単語の組み合わせ
        st.subheader("Combinaciones de palabras observadas con cierta frecuencia")
        bigram_freq = Counter(list(zip(filtered_words[:-1], filtered_words[1:])))
        trigram_freq = Counter(list(zip(filtered_words[:-2], filtered_words[1:-1], filtered_words[2:])))
        
        col1, col2 = st.columns(2)
        with col1:
            st.write("##### :blue[Bigrams observados:]")
            for bigram, freq in bigram_freq.most_common(3):
                st.write(f"{' '.join(bigram)}: {freq}")
        with col2:
            st.write("##### :blue[Trigrams observados:]")
            for trigram, freq in trigram_freq.most_common(3):
                st.write(f"{' '.join(trigram)}: {freq}")

